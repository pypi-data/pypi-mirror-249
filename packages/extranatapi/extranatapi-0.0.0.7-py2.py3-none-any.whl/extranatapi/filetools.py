# -*- coding: utf-8 -*-
"""File tools."""

import csv
import logging
import os
import sys

from openpyxl import load_workbook, Workbook
from .base import ExtranatObjectCollection
from .nageur import Nageur


DEFAULT_CSV_SEPARATOR = ';'
DEFAULT_EOL = '\n'
HEADER_NAGEUR = ['nageur', 'annee', 'age', 'iuf', 'sexe', 'bassin', 'nage', 'temps']
HEADER_IUF = ['iuf']
HEADER_EQUIPE_NAGES = ['nage']
HEADER_EQUIPE_NAGEUR = ['nageur', 'annee', 'age', 'iuf', 'sexe', 'bassin', 'nage', 'temps', 'points']


class FileTools():
    """File tools."""

    def __init__(self, logger=None):
        """Construtor."""
        if logger is None:
            self.logger = logging.getLogger()
        else:
            self.logger = logger

    @staticmethod
    def __excel_sheet_to_dict__(sheet):
        """Excel sheet to dict"""
        results = []
        row_count = sheet.max_row
        column_count = sheet.max_column
        header = []

        for col in range(column_count):
            header.append(sheet.cell(1, col + 1).value.upper())

        for row in range(1, row_count):
            dictvalues = {}
            for col in range(column_count):
                dictvalues[header[col]] = sheet.cell(row + 1, col + 1).value
            results.append(dictvalues)

        return results

    def __check_nageurs_header__(self, header, keys):
        """Check nageurs header."""
        result = True
        result_keys = {}

        for item_key in keys:

            found_item_key = False
            for item_header in header:
                if item_key.lower() == item_header.lower():
                    found_item_key = True
                    result_keys[item_key] = item_header

            if not found_item_key:
                self.logger.critical('%s non trouvé dans les colonnes', item_key)
                result = False

        return result, result_keys

    def read_uif(self, filename):
        """Read iuf in file."""
        if filename is None or not os.path.exists(filename):
            self.logger.critical('File "%s" does\'nt exist', filename)
            sys.exit(-1)

        _, file_extension = os.path.splitext(filename)
        file_extension = file_extension.lower()

        if file_extension == '.csv':
            return self.__read_uif_csv__(filename)
        if file_extension == '.xlsx':
            return self.__read_uif_excel__(filename)

        self.logger.critical('Extension "%s" unknown', file_extension)
        sys.exit(-1)

    @staticmethod
    def __get_encoding_type__(current_file):
        """Get Encoding Type."""
        result = 'cp437'

        try:
            with open(current_file, 'r', encoding='utf-8') as filetoread:
                filetoread.read()
                result = 'utf-8'
        except UnicodeDecodeError:
            pass

        return result

    def read_nageurs(self, filename):
        """Read nageurs in file."""
        if filename is None or not os.path.exists(filename):
            self.logger.critical('File "%s" does\'nt exist', filename)
            sys.exit(-1)

        _, file_extension = os.path.splitext(filename)
        file_extension = file_extension.lower()

        if file_extension == '.csv':
            return self.__read_nageurs_csv__(filename)
        if file_extension == '.xlsx':
            return self.__read_nageurs_excel__(filename)

        self.logger.critical('Extension "%s" unknown', file_extension)
        sys.exit(-1)

    def read_equipe_nageurs(self, filename):
        """Read nageurs in file."""
        if filename is None or not os.path.exists(filename):
            self.logger.critical('File "%s" does\'nt exist', filename)
            sys.exit(-1)

        _, file_extension = os.path.splitext(filename)
        file_extension = file_extension.lower()

        if file_extension == '.csv':
            return self.__read_equipe_nageurs_csv__(filename)
        if file_extension == '.xlsx':
            return self.__read_equipe_nageurs_excel__(filename)

        self.logger.critical('Extension "%s" unknown', file_extension)
        sys.exit(-1)

    def read_equipe_nages(self, filename):
        """Read nages for a Team in file."""
        if filename is None or not os.path.exists(filename):
            self.logger.critical('File "%s" does\'nt exist', filename)
            sys.exit(-1)

        _, file_extension = os.path.splitext(filename)
        file_extension = file_extension.lower()

        if file_extension == '.csv':
            return self.__read_equipe_nages_csv__(filename)
        if file_extension == '.xlsx':
            return self.__read_equipe_nages_excel__(filename)

        self.logger.critical('Extension "%s" unknown', file_extension)
        sys.exit(-1)

    def __read_uif_csv__(self, filename):
        """Read iuf in CSV file."""
        self.logger.debug('Read %s', filename)

        results = []

        inputformat = self.__get_encoding_type__(filename)
        with open(filename, encoding=inputformat) as csvfile:

            items = csv.DictReader(csvfile, delimiter=';')
            check, header = self.__check_nageurs_header__(header=items.fieldnames, keys=HEADER_IUF)
            if not check:
                return results

            for item in items:

                iuf = item[header['iuf']]
                results.append(iuf)

        return results

    def __read_uif_excel__(self, filename):
        """Read iuf in Excel file."""
        self.logger.debug('Read %s', filename)
        results = []

        # Opening Excel file
        current_workbook = load_workbook(filename=filename)
        sheet = current_workbook.worksheets[0]

        items = self.__excel_sheet_to_dict__(sheet)

        check = False
        header = None
        if len(items) > 0:
            check, header = self.__check_nageurs_header__(header=items[0], keys=HEADER_IUF)
        if not check:
            return results

        for item in items:

            iuf = item[header['iuf']]
            results.append(iuf)

        return results

    def __read_nageurs_csv__(self, filename):
        """Read nageurs in CSV file."""
        self.logger.debug('Read %s', filename)
        results = ExtranatObjectCollection(Nageur())
        tmp_result = {}

        inputformat = self.__get_encoding_type__(filename)
        with open(filename, encoding=inputformat) as csvfile:

            items = csv.DictReader(csvfile, delimiter=';')
            header = None
            check, header = self.__check_nageurs_header__(header=items.fieldnames, keys=HEADER_NAGEUR)
            if not check:
                return results

            for item in items:

                # get nageur
                iuf = item[header['iuf']]
                if iuf not in tmp_result:
                    tmp_result[iuf] = Nageur(
                        iuf=item[header['iuf']],
                        name=item[header['nageur']],
                        sexe=item[header['sexe']],
                        yearofbirth=item[header['annee']],
                        age=item[header['age']])

                # get nage
                tmp_result[iuf].addnage(
                    nage=item[header['nage']],
                    temps=item[header['temps']],
                    bassin=item[header['bassin']])

        for _, value in tmp_result.items():
            results.append(value)

        return results

    def __read_nageurs_excel__(self, filename):
        """Read nageurs in Excel file."""
        self.logger.debug('Read %s', filename)
        results = ExtranatObjectCollection(Nageur())

        # Opening Excel file
        current_workbook = load_workbook(filename=filename)
        sheet = current_workbook.worksheets[0]

        items = self.__excel_sheet_to_dict__(sheet)

        check = False
        header = None
        if len(items) > 0:
            check, header = self.__check_nageurs_header__(header=items[0], keys=HEADER_NAGEUR)
        if not check:
            return results

        tmp_result = {}
        for item in items:

            # get nageur
            iuf = item[header['iuf']]
            if iuf not in tmp_result:
                tmp_result[iuf] = Nageur(
                    iuf=item[header['iuf']],
                    name=item[header['nageur']],
                    sexe=item[header['sexe']],
                    yearofbirth=item[header['annee']],
                    age=item[header['age']])

            # get nage
            tmp_result[iuf].addnage(
                nage=item[header['nage']],
                temps=item[header['temps']],
                bassin=item[header['bassin']])

        for _, value in tmp_result.items():
            results.append(value)

        return results

    def __read_equipe_nages_csv__(self, filename):
        """Read nages for a Team in CSV file."""
        self.logger.debug('Read %s', filename)
        results = []

        inputformat = self.__get_encoding_type__(filename)
        with open(filename, encoding=inputformat) as csvfile:

            items = csv.DictReader(csvfile, delimiter=';')
            check, header = self.__check_nageurs_header__(header=items.fieldnames, keys=HEADER_EQUIPE_NAGES)
            if not check:
                return results

            for item in items:
                results.append(item[header['nage']])

        return results

    def __read_equipe_nages_excel__(self, filename):
        """Read nages for a Team in Excel file."""
        self.logger.debug('Read %s', filename)
        results = []

        # Opening Excel file
        current_workbook = load_workbook(filename=filename)
        sheet = current_workbook.worksheets[0]

        items = self.__excel_sheet_to_dict__(sheet)

        check = False
        header = None
        if len(items) > 0:
            check, header = self.__check_nageurs_header__(header=items[0], keys=HEADER_EQUIPE_NAGES)
        if not check:
            return results

        for item in items:
            results.append(item[header['nage']])

        return results

    def __read_equipe_nageurs_csv__(self, filename):
        """Read nageurs in CSV file."""
        self.logger.debug('Read %s', filename)
        results = ExtranatObjectCollection(Nageur())
        tmp_result = {}

        inputformat = self.__get_encoding_type__(filename)
        with open(filename, encoding=inputformat) as csvfile:

            items = csv.DictReader(csvfile, delimiter=';')
            check, header = self.__check_nageurs_header__(header=items.fieldnames, keys=HEADER_EQUIPE_NAGEUR)
            if not check:
                return results

            for item in items:

                # get nageur
                iuf = item[header['iuf']]
                if iuf not in tmp_result:
                    tmp_result[iuf] = Nageur(
                        iuf=item[header['iuf']],
                        name=item[header['nageur']],
                        sexe=item[header['sexe']],
                        yearofbirth=item[header['annee']],
                        age=item[header['age']])

                # get nage
                tmp_result[iuf].addnage(
                    nage=item[header['nage']],
                    temps=item[header['temps']],
                    bassin=item[header['bassin']],
                    points=int(item[header['points']].replace(' pts', '')))

        for _, value in tmp_result.items():
            results.append(value)

        return results

    def __read_equipe_nageurs_excel__(self, filename):
        """Read nageurs in Excel file."""
        self.logger.debug('Read %s', filename)
        results = ExtranatObjectCollection(Nageur())

        # Opening Excel file
        current_workbook = load_workbook(filename=filename)
        sheet = current_workbook.worksheets[0]

        items = self.__excel_sheet_to_dict__(sheet)

        check = False
        header = None
        if len(items) > 0:
            check, header = self.__check_nageurs_header__(header=items[0], keys=HEADER_EQUIPE_NAGEUR)
        if not check:
            return results

        tmp_result = {}
        for item in items:

            # get nageur
            iuf = item[header['iuf']]
            if iuf not in tmp_result:
                tmp_result[iuf] = Nageur(
                    iuf=item[header['iuf']],
                    name=item[header['nageur']],
                    sexe=item[header['sexe']],
                    yearofbirth=item[header['annee']],
                    age=item[header['age']])

            # get nage
            tmp_result[iuf].addnage(
                nage=item[header['nage']],
                temps=item[header['temps']],
                bassin=item[header['bassin']],
                points=int(item[header['points']].replace(' pts', '')))

        for _, value in tmp_result.items():
            results.append(value)

        return results

    def write_csv(self, array, array_header, output=sys.stdout, separator=DEFAULT_CSV_SEPARATOR, endofline=DEFAULT_EOL):  # pylint: disable=too-many-arguments
        """Write data in CSV format."""
        self.logger.debug('Write CVS')

        if len(array) > 0:
            if not isinstance(array[0], list):
                array = [array]

        output.write(separator.join(array_header) + endofline)

        for item in array:
            output.write(separator.join(item) + endofline)

    def write_excel(self, filename, array, array_header):
        """Write data in Excel file."""
        self.logger.debug('Write %s', filename)
        if len(array) > 0:
            if not isinstance(array[0], list):
                array = [array]

        # Create Excel Workbook / sheet
        current_workbook = Workbook()
        current_worksheet = current_workbook.active
        current_worksheet.title = 'info'

        # create headers
        i = 0
        num_row = 1
        for item in array_header:

            current_worksheet.cell(row=num_row, column=(i + 1), value=item)
            i += 1

        # loop on rows
        for item in array:
            num_row += 1
            col = 1
            for val in item:
                current_worksheet.cell(row=num_row, column=col, value=val)
                col = col + 1

        # Save
        current_workbook.save(filename)
