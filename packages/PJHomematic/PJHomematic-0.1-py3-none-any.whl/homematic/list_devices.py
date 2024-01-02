import sys
import time
from collections import namedtuple

OPTION_OPENPYXL = False
try:
    import openpyxl
    from openpyxl.styles import Font

    OPTION_OPENPYXL = True
except ImportError:
    pass

from homematic import hmrpc

COLUMN_NAMES = 'Name', 'Type', 'Address', 'Subsection', 'Room', 'Level'


def get_devices(url, user, passwd, header=False):
    hmrc = hmrpc.HMRPCClient(url, username=user, password=passwd)

    # Get room info
    rooms = hmrc.call('Room.getAll')
    dev_rooms = {}
    for r in rooms:
        for chid in r['channelIds']:
            if chid in dev_rooms:
                dev_rooms[chid] += [r['name']]
            else:
                dev_rooms[chid] = [r['name']]

    # Get subsection info
    subs = hmrc.call('Subsection.getAll')
    dev_subs = {}
    for s in subs:
        for chid in s['channelIds']:
            if chid in dev_subs:
                dev_subs[chid] += [s['name']]
            else:
                dev_subs[chid] = [s['name']]

    # Get device info and merge with subsections and rooms
    details = hmrc.call('Device.listAllDetail')

    Device = namedtuple('Device', COLUMN_NAMES)

    if header:
        yield Device(*COLUMN_NAMES)

    for d in details:
        rooms = []
        if d['id'] in dev_rooms:
            rooms = dev_rooms[d['id']]

        subs = []
        if d['id'] in dev_subs:
            subs = dev_subs[d['id']]

        yield Device(d['name'], d['type'], d['address'], ','.join(subs), ','.join(rooms), 'Device')

        for c in d['channels']:
            rooms = []
            if c['id'] in dev_rooms:
                rooms = dev_rooms[c['id']]

            subs = []
            if c['id'] in dev_subs:
                subs = dev_subs[c['id']]

            yield Device(c['name'], d['type'], c['address'], ','.join(subs), ','.join(rooms), 'Channel')


def print_list(url, user, passwd, header=True):
    for d in get_devices(url, user, passwd, header):
        print('\t'.join(d))


def export_excel(url, user, passwd, xl_file=time.strftime('%Y-%m-%d_%H.%M.%S') + '_HM-Devices.xlsx', header=True):
    if not OPTION_OPENPYXL:
        sys.exit('Missing module openpyxl!')

    xl_wb = openpyxl.Workbook()
    xl_ws = xl_wb.active
    xl_ws.title = 'HM-Devices'

    if header:
        xl_ws.freeze_panes = 'A2'

    row = 1
    columns = 0
    col_widths = [0] * len(COLUMN_NAMES)
    for d in get_devices(url, user, passwd, header):
        column = 1
        for val in d:
            c = xl_ws.cell(column=column, row=row, value=val)

            val_len = len(str(val))
            if col_widths[column - 1] < val_len:
                col_widths[column - 1] = val_len

            if header and row == 1:
                columns = len(d)
                c.font = Font(bold=True)

            column += 1

        row += 1

    if header:
        xl_ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(columns)}1'

    # Fit size to content width approximation
    for c, w in zip(range(1, len(col_widths)+1), col_widths):
        # Add 5 due to Excel filter drop down
        xl_ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w + 5

    try:
        xl_wb.save(xl_file)
    except OSError as e:
        sys.exit(str(e))


if __name__ == '__main__':
    if len(sys.argv) > 3:
        addr = sys.argv[1]
        user = sys.argv[2]
        pwd = sys.argv[3]
    else:
        sys.exit(f'Try: {sys.argv[0]} ADDRESS USERNAME PASSWORD')

    if OPTION_OPENPYXL:
        export_excel(addr, user, pwd)
    else:
        print_list(addr, user, pwd)
