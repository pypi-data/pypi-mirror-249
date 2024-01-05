#!/usr/bin/metrics_db_name python
# -*- coding: utf-8 -*-
# @Time:2023/1/6 11:23
# @Author:boyizhang
from collections import defaultdict

import openpyxl


class ExcelTools():
    @staticmethod
    def write_to_file(sheet_data_dict, headers, filename=None, type=1):
        """
        :param sheet_data_dict: defaultdict 每个key对应一个sheet的值。sheet值的key的写入数据与headers元素顺序一致
        :param headers: list 标题头
        :param filename: 保存的文件地址+文件名
        :param type:
        :return:
        """
        workbook = openpyxl.Workbook()
        i = 0

        for sheet_name, sheet_data in sheet_data_dict.items():
            if i == 0:
                worksheet = workbook.worksheets[0]
            else:
                worksheet = workbook.create_sheet()
            worksheet.title = sheet_name
            # for i in enumerate(headers):
            #     worksheet.(i + 1, 1, data)
            # for i, data in enumerate(sheet_data.get('shop_ids')):
            #     worksheet.cell(i + 1, 1, data)
            # for i, data in enumerate(sheet_data.get('listing_limit')):
            #     worksheet.cell(i + 1, 2, data)
            # for i, data in enumerate(sheet_data.get('shop_ids')):
            #     worksheet.cell(i + 1, 3, data)
            if headers:
                worksheet.append(headers)
            min_len = 0
            for item_list in sheet_data.values():
                if min_len != 0 and min_len > len(item_list) or min_len == 0:
                    min_len = len(item_list)

            for i in range(min_len):
                row_list = []
                for item_list in sheet_data.values():
                    row_list.append(item_list[i])
                worksheet.append(row_list)

            # 方法二
            # shop_ids = sheet_data.get('shop_ids')
            # remaks = sheet_data.get('remaks')
            # if type == 1:
            #     listing_limit = sheet_data.get('listing_limit')
            #     for i in range(len(shop_ids)):
            #         worksheet.append([shop_ids[i], listing_limit[i], remaks[i]])
            # elif type == 2:
            #     start_date = sheet_data.get('start_date')
            #     end_date = sheet_data.get('end_date')
            #     for i in range(len(shop_ids)):
            #         worksheet.append([shop_ids[i], start_date[i], end_date[i], remaks[i]])
            # elif type == 3:
            #     category_ids = sheet_data.get('category_ids')
            #     for i in range(len(shop_ids)):
            #         worksheet.append([shop_ids[i], category_ids[i], remaks[i]])
            i += 1
        workbook.save(filename=filename)

    @staticmethod
    def read_data(filename):
        workbook = openpyxl.load_workbook(filename)
        sheet_names = workbook.sheetnames
        for sheet_name in sheet_names:
            # 获取工作表对象
            worksheet = workbook[sheet_name]
            name = worksheet.title
            rows = worksheet.max_row
            columns = worksheet.max_column
            print(sheet_name, worksheet, name, rows, columns, worksheet.rows, worksheet.columns)

            # 按行读取
            for row in worksheet.rows:
                for cell in row:
                    print(cell.value, end=' ')
                print()

            # 按列读取
            # for col in worksheet.columns:
            #     for cell in col:
            #         print(cell.value, end=' ')  #
            # print()
            # print("-"*100)

            # 读取指定行数, 指定的列
            # for  rows in  list(worksheet.rows)[0:3]:
            #     for cell in rows[0:3]:
            #         print(cell.value, end=" ")
            #     print()  # 读取指定行数, 指定的列
            # for i in range(1, 4):
            #     for j in range(1, 4):
            #         print(worksheet.cell(row=i, column=j).value, end=" ")
            #     print()

            # 插入一列数据

    @staticmethod
    def update_sheet(filename):
        workbook = openpyxl.load_workbook(filename=filename)
        worksheet = workbook.worksheets[0]
        # 在第一列之前插入一列
        # worksheet.insert_cols(1)
        # for index,row in enumerate(worksheet.rows):
        #     if index == 0:
        #         row[0].value = "测试"
        #     else:
        #         row[0].value = str(index)+'- test'

        worksheet.delete_cols(1, 2)
        # worksheet.delete_rows(0,3)
        workbook.save(filename)

    @staticmethod
    def call_case():
        filename = 'white_list-100000_3.xlsx'
        headers = ["Shop ID", "Exemption type", "Exemption start date(YYYY/MM/DD)", "Exemption end date(YYYY/MM/DD)",
                   "Remarks(Character limit 1000)"]
        sheet_data_dict = defaultdict(dict)
        sheet_name = 'sheet'
        shop_ids = [i * 10 for i in range(100000, 100000 + 99988)]
        exemption_type = ["Send" for i in range(len(shop_ids))]
        exemption_start_date = ["2023/02/01" for i in range(len(shop_ids))]
        exemption_end_date = ["2023/02/05" for i in range(len(shop_ids))]
        remaks = [str(shop_id) + "-remarks-autogenerate" for shop_id in shop_ids]
        init_shop_id = shop_ids[-1] // 10 + 10 * 1
        # 按照headers元素的顺序，给其对应的key赋值
        sheet_data_dict[sheet_name]['shop_ids'] = shop_ids
        sheet_data_dict[sheet_name]['exemption_type'] = exemption_type
        sheet_data_dict[sheet_name]['exemption_start_date'] = exemption_start_date
        sheet_data_dict[sheet_name]['exemption_end_date'] = exemption_end_date
        sheet_data_dict[sheet_name]['remaks'] = remaks

        ExcelTools.write_to_file(sheet_data_dict, headers, filename)  # ExcelTools.read_data(filename)


if __name__ == '__main__':
    filename = 'white_list-100000_3.xlsx'
    headers = ["Shop ID", "Exemption type", "Exemption start date(YYYY/MM/DD)", "Exemption end date(YYYY/MM/DD)",
               "Remarks(Character limit 1000)"]
    sheet_data_dict = defaultdict(dict)
    sheet_name = 'sheet'
    shop_ids = [i * 10 for i in range(100000, 100000 + 99988)]
    exemption_type = ["Send" for i in range(len(shop_ids))]
    exemption_start_date = ["2023/02/01" for i in range(len(shop_ids))]
    exemption_end_date = ["2023/02/05" for i in range(len(shop_ids))]
    remaks = [str(shop_id) + "-remarks-autogenerate" for shop_id in shop_ids]
    init_shop_id = shop_ids[-1] // 10 + 10 * 1
    # 按照headers元素的顺序，给其对应的key赋值
    sheet_data_dict[sheet_name]['shop_ids'] = shop_ids
    sheet_data_dict[sheet_name]['exemption_type'] = exemption_type
    sheet_data_dict[sheet_name]['exemption_start_date'] = exemption_start_date
    sheet_data_dict[sheet_name]['exemption_end_date'] = exemption_end_date
    sheet_data_dict[sheet_name]['remaks'] = remaks

    ExcelTools.write_to_file(sheet_data_dict, headers, filename)  # ExcelTools.read_data(filename)
